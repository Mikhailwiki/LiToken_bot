import io
import qrcode
import pymorphy3
import csv

from openpyxl import Workbook


async def make_qrcode(to_chat_id: int, amount: int):
    url = f'https://t.me/LiToken_bot?start=send_{to_chat_id}_{amount}'
    qrcode_img = qrcode.make(url)
    buf = io.BytesIO()
    qrcode_img.save(buf)
    buf.seek(0)
    return buf, url


async def agree_with_num(word: str, number: int) -> str:
    morph = pymorphy3.MorphAnalyzer()
    return morph.parse(word)[0].make_agree_with_number(number).word


async def load_data(filename):
    data = []
    with open(filename, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            data.append(row)
    return data


async def find_user_by_username(username, data):
    for user in data:
        if user['телеграм'] == username:
            return {
                'ФИ': user['ФИ'],
                'Класс': user['Класс']
            }
    return None


async def get_table(table) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Пользователи и балансы'
    data = await load_data('students.csv')

    for row, (coins, username) in enumerate(table):
        row += 1
        student = await find_user_by_username(f'@{username}', data)
        if student:
            name, class_ = student['ФИ'], student['Класс']
            ws[f'A{row}'] = name
            ws[f'B{row}'] = class_
        else:
            ws[f'A{row}'] = '_'
            ws[f'B{row}'] = '_'
        ws[f'C{row}'] = username
        ws[f'D{row}'] = str(coins)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 20


    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
